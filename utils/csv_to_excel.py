# utils/csv_to_excel.py
"""
CSV转Excel工具函数
"""
import pandas as pd
import chardet
import os
from datetime import datetime
from openpyxl.utils import get_column_letter


def detect_encoding(file_path):
    """检测文件编码"""
    try:
        with open(file_path, 'rb') as f:
            raw_data = f.read(10000)
            result = chardet.detect(raw_data)
            return result['encoding']
    except Exception as e:
        print(f"编码检测失败: {e}")
        return None


def get_display_length(text):
    """获取文本的显示长度（考虑中文字符）"""
    if not text or pd.isna(text):
        return 0
    text_str = str(text)
    length = 0
    for char in text_str:
        if '\u4e00' <= char <= '\u9fff':
            length += 2
        else:
            length += 1
    return length


def get_csv_max_column_widths(file_path, encoding, sample_lines=100):
    """获取CSV文件每列的最大显示宽度"""
    max_widths = {}
    try:
        if encoding == "auto":
            detected_encoding = detect_encoding(file_path)
            if detected_encoding:
                encoding = detected_encoding
            else:
                encoding = 'utf-8'

        with open(file_path, 'r', encoding=encoding, errors='replace') as f:
            first_line = f.readline().strip()
            headers = first_line.split(',')

            for i, header in enumerate(headers):
                if header.startswith('Unnamed:'):
                    continue
                header_display_len = get_display_length(header)
                max_widths[i] = header_display_len

            lines_read = 0
            for line in f:
                if lines_read >= sample_lines:
                    break
                values = line.strip().split(',')
                for i, value in enumerate(values):
                    if i < len(headers):
                        if i < len(headers) and headers[i].startswith('Unnamed:'):
                            continue
                        value_display_len = get_display_length(value)
                        if value_display_len > max_widths.get(i, 0):
                            max_widths[i] = value_display_len
                lines_read += 1
        return max_widths
    except Exception as e:
        print(f"获取CSV列宽失败: {e}")
        return None


def calculate_column_width(df, csv_widths=None, width_setting="自动调整", fixed_width=15):
    """计算Excel列宽"""
    column_widths = {}

    for idx, col_name in enumerate(df.columns):
        default_width = 12

        if width_setting == "固定宽度":
            width = float(fixed_width)
        else:
            header_len = get_display_length(col_name)

            if width_setting == "紧凑宽度":
                width = max(header_len * 0.9, 8)
            else:  # 自动调整
                if csv_widths and idx in csv_widths:
                    csv_width = csv_widths[idx]
                    width = csv_width * 0.9
                else:
                    max_data_len = 0
                    sample_size = min(len(df), 100)
                    for i in range(sample_size):
                        value = df.iloc[i, idx]
                        if pd.notna(value):
                            value_len = get_display_length(str(value))
                            if value_len > max_data_len:
                                max_data_len = value_len
                    max_len = max(header_len, max_data_len)
                    width = max_len * 0.9

            width = max(8, min(width, 50))

        column_widths[idx] = width

    return column_widths


def apply_column_widths(writer, sheet_name, df, column_widths):
    """应用列宽设置"""
    worksheet = writer.sheets[sheet_name]
    for idx, width in column_widths.items():
        if idx < len(df.columns):
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = width


def read_csv_with_encoding(file_path, encoding="auto"):
    """使用指定编码读取CSV文件"""
    try:
        if encoding == "auto":
            detected_encoding = detect_encoding(file_path)
            if detected_encoding:
                encoding = detected_encoding
            else:
                encoding = 'utf-8'

        df = pd.read_csv(file_path, encoding=encoding, on_bad_lines='warn', dtype=object)
        return df, encoding
    except Exception as e:
        print(f"使用 {encoding} 编码读取失败: {e}")
        return None, None


def csv_to_excel(csv_path, output_path=None, encoding="auto",
                 num_format="文本格式", null_display="NULL",
                 width_setting="自动调整", fixed_width=15,
                 clean_empty=True):
    """
    将CSV文件转换为Excel文件

    Args:
        csv_path: CSV文件路径
        output_path: 输出Excel路径（None则自动生成）
        encoding: 编码方式
        num_format: 数字格式（"文本格式"/"数字格式"）
        null_display: 空值显示（"NULL"/"空"/"原始值"）
        width_setting: 列宽设置（"自动调整"/"紧凑宽度"/"固定宽度"）
        fixed_width: 固定宽度值
        clean_empty: 是否清理Unnamed列

    Returns:
        (success, output_path, message)
    """
    try:
        # 获取CSV文件的列宽信息
        csv_widths = None
        if width_setting == "自动调整":
            csv_widths = get_csv_max_column_widths(csv_path, encoding)

        # 尝试读取文件
        df, used_encoding = read_csv_with_encoding(csv_path, encoding)

        if df is None:
            # 尝试其他编码
            common_encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'latin1', 'cp1252']
            for enc in common_encodings:
                if enc == encoding and encoding != "auto":
                    continue
                df, used_encoding = read_csv_with_encoding(csv_path, enc)
                if df is not None:
                    break

            if df is None:
                return False, None, "无法读取CSV文件，请尝试手动指定编码"

        # 清理Unnamed列
        if clean_empty:
            columns_to_drop = [col for col in df.columns if str(col).startswith('Unnamed:')]
            if columns_to_drop:
                df = df.drop(columns=columns_to_drop, errors='ignore')

        # 处理空值
        if null_display == "NULL":
            df = df.fillna('NULL')
        elif null_display == "空":
            df = df.fillna('')

        # 生成输出文件名
        if output_path is None:
            base_name = os.path.splitext(csv_path)[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"{base_name}_{timestamp}.xlsx"

        # 计算列宽
        column_widths = calculate_column_width(df, csv_widths, width_setting, fixed_width)

        # 保存为Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # 处理长数字列
            for col_idx, col_name in enumerate(df.columns):
                col_letter = get_column_letter(col_idx + 1)

                # 检查是否有长数字
                has_long_numbers = False
                sample_size = min(len(df), 100)
                for i in range(sample_size):
                    cell = df.iloc[i, col_idx]
                    if pd.notna(cell):
                        cell_str = str(cell).strip()
                        if cell_str.isdigit() and len(cell_str) > 15:
                            has_long_numbers = True
                            break

                if has_long_numbers and num_format == "文本格式":
                    for row_idx in range(2, len(df) + 2):
                        cell = worksheet[f"{col_letter}{row_idx}"]
                        if pd.notna(df.iloc[row_idx - 2, col_idx]):
                            cell.value = str(df.iloc[row_idx - 2, col_idx])
                            cell.number_format = '@'

            # 应用列宽
            apply_column_widths(writer, 'Sheet1', df, column_widths)

        return True, output_path, f"成功转换: {os.path.basename(csv_path)} -> {os.path.basename(output_path)}"

    except Exception as e:
        return False, None, f"转换失败: {str(e)}"