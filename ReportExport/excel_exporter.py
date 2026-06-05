import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color='4361ee', end_color='4361ee', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name='微软雅黑', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=False)


def _style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_body(ws, row, col_count, align=CENTER):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.alignment = align
        cell.border = THIN_BORDER


def _recolor_header_firstcol(ws, col_count):
    BLUE2_FILL = PatternFill(start_color='42A5F5', end_color='42A5F5', fill_type='solid')
    BLUE1_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    BLACK_BOLD_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = BLUE2_FILL
        cell.font = BLACK_BOLD_FONT
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).fill = BLUE1_FILL


def _recolor_header_firstcol_green(ws, col_count):
    GREEN2_FILL = PatternFill(start_color='66BB6A', end_color='66BB6A', fill_type='solid')
    GREEN1_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    BLACK_BOLD_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = GREEN2_FILL
        cell.font = BLACK_BOLD_FONT
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).fill = GREEN1_FILL


def _write_sheet(ws, title, headers, data_rows, col_widths=None):
    ws.title = title
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, len(headers))
    for ri, row in enumerate(data_rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=str(val))
        _style_body(ws, ri, len(headers))
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + ci)].width = w
    else:
        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + ci)].width = max(12, len(str(headers[ci - 1])) * 2)


def build_excel(data, note_data, week_label='', is_monthly=False):
    wb = Workbook()
    ws_index = 0

    type_name = '月报' if is_monthly else '周报'
    title_row = f'媒介-审计报告（{week_label}）'

    # ── 总体分析 ──
    if data:
        from ReportExport.report_generator import (
            build_summary_table_data, build_workload_detail_data,
            build_quality_detail_data, build_cost_performance_data,
            build_rebate_analysis_data, build_level_analysis_data,
            build_cost_analysis_data,
        )

        # Sheet 1: 总结
        ws_index += 1
        ws = wb.active if ws_index == 1 else wb.create_sheet()
        table_rows = build_summary_table_data(data)
        if table_rows:
            headers = ['所属小组', '媒介人数', '签框量(口头)', '签框量(书面)', '总提报量', '总定档量', '平均定档数', '过筛率中位数', '平均定档成本']
            data_rows = []
            for r in table_rows:
                cost = f"{r['avg_cost']:.2f}" if isinstance(r['avg_cost'], (int, float)) else str(r['avg_cost'])
                data_rows.append([r['group'], r['media_count'], r['kou'], r['shu'], r['tibao'], r['dingdang'], r['avg_dingdang'], r['guoshuai'], cost])
            _write_sheet(ws, '总结', headers, data_rows)

        # Sheet 2: 工作量
        ws_index += 1
        ws = wb.create_sheet()
        workload = build_workload_detail_data(data)
        if workload:
            headers = ['小组', '姓名', '提报数', '定档数']
            data_rows = []
            for g in workload:
                for p in g['people']:
                    data_rows.append([g['group'], p['name'], str(p['tibao']), str(p['dingdang'])])
            _write_sheet(ws, '工作量', headers, data_rows)

        # Sheet 3: 工作质量
        ws_index += 1
        ws = wb.create_sheet()
        premium = build_quality_detail_data(data, 'premium')
        data_rows = []
        if premium:
            for p in premium:
                data_rows.append([p['name'], p['group'], str(p['total']), str(p['passed']), p['rate_str'], p['evaluation']])
        high_read = build_quality_detail_data(data, 'high_read')
        if high_read:
            for p in high_read:
                data_rows.append([p['name'], p['group'], str(p['total']), str(p['passed']), p['rate_str'], p['evaluation']])
        if data_rows:
            _write_sheet(ws, '工作质量', ['对应名字', '所属小组', '总提报达人数', '过筛人数', '过筛率(%)', '质量评估'], data_rows)
            RED_FILL = PatternFill(start_color='EF5350', end_color='EF5350', fill_type='solid')
            RED_LIGHT_FILL = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            BLACK_BOLD_FONT = Font(name='微软雅黑', bold=True, color='000000', size=10)
            for c in range(1, 7):
                cell = ws.cell(row=1, column=c)
                cell.fill = RED_FILL
                cell.font = BLACK_BOLD_FONT
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=1).fill = RED_LIGHT_FILL

        # Sheet 4: 成本发挥
        ws_index += 1
        ws = wb.create_sheet()
        cost_perf = build_cost_performance_data(data)
        if cost_perf:
            headers = ['媒介组', '定档达人数', '平均返点比例(%)', '返点比例最高', '返点比例最低']
            data_rows = [[r['group'], str(r['dingdang']), r['avg_rebate'], r['high'], r['low']] for r in cost_perf]
            _write_sheet(ws, '成本发挥', headers, data_rows)
            BLUE2_FILL = PatternFill(start_color='42A5F5', end_color='42A5F5', fill_type='solid')
            BLUE1_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            BLACK_BOLD_FONT = Font(name='微软雅黑', bold=True, color='000000', size=10)
            for c in range(1, 6):
                cell = ws.cell(row=1, column=c)
                cell.fill = BLUE2_FILL
                cell.font = BLACK_BOLD_FONT
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=1).fill = BLUE1_FILL

        # Sheet 5: 返点分析
        ws_index += 1
        ws = wb.create_sheet()
        rebate_rows = build_rebate_analysis_data(data)
        if rebate_rows:
            headers = ['定档媒介', '定档达人数', '所属小组', '平均返点比例(%)', '返点比例最大值(%)', '返点比例最小值(%)', '返点比例中位数(%)',
                       '总返点金额(元)', '平均返点金额(元)', '返点金额最大值(元)', '返点金额最小值(元)', '返点金额中位数(元)',
                       '返点表现评估', '返点优化建议']
            data_rows = []
            for r in rebate_rows:
                data_rows.append([r['name'], str(r['dingdang']), r['group'], r['avg_rebate'], r['max_rebate'], r['min_rebate'],
                                  r['median_rebate'], f"{r['total_rebate_amt']:.2f}", f"{r['avg_rebate_amt']:.2f}",
                                  f"{r['max_rebate_amt']:.2f}", f"{r['min_rebate_amt']:.2f}", f"{r['median_rebate_amt']:.2f}",
                                  r['evaluation'], r['suggestion']])
            _write_sheet(ws, '返点分析', headers, data_rows)
            _recolor_header_firstcol(ws, 14)

        # Sheet 6: 达人量级分析
        ws_index += 1
        ws = wb.create_sheet()
        lvl_rows = build_level_analysis_data(data)
        if lvl_rows:
            headers = ['定档媒介', '达人量级', '达人数', '所属小组', '总成本(元)', '平均成本(元)', '总返点金额(元)', '平均返点金额(元)',
                       '平均返点比例(%)', '总互动量', '平均互动量', '总阅读量', '平均阅读量', '平均CPE', '平均CPM']
            data_rows = []
            for r in lvl_rows:
                data_rows.append([r['name'], r['level'], str(r['count']), r['group'], f"{r['total_cost']:.2f}", f"{r['avg_cost']:.2f}",
                                  f"{r['total_rebate']:.2f}", f"{r['avg_rebate']:.2f}", r['avg_rebate_pct'],
                                  f"{r['total_interact']:.0f}", f"{r['avg_interact']:.2f}", f"{r['total_read']:.0f}",
                                  f"{r['avg_read']:.2f}", f"{r['avg_cpe']:.2f}", f"{r['avg_cpm']:.2f}"])
            _write_sheet(ws, '达人量级分析', headers, data_rows)
            _recolor_header_firstcol(ws, 15)

        # Sheet 7: 成本分析
        ws_index += 1
        ws = wb.create_sheet()
        cost_rows = build_cost_analysis_data(data)
        if cost_rows:
            headers = ['定档媒介', '定档达人数', '所属小组', '总成本(元)', '平均成本(元)', '成本最大值(元)', '成本最小值(元)', '成本中位数(元)',
                       '总报价(元)', '平均报价(元)', '报价最大值(元)', '报价最小值(元)', '总下单价(元)', '平均下单价(元)',
                       '总节约金额(元)', '平均节约金额(元)', '成本占比(%)', '总返点金额(元)', '平均返点金额(元)']
            data_rows = []
            for r in cost_rows:
                data_rows.append([r['name'], str(r['count']), r['group'], f"{r['total_cost']:.2f}", f"{r['avg_cost']:.2f}",
                                  f"{r['max_cost']:.2f}", f"{r['min_cost']:.2f}", f"{r['median_cost']:.2f}",
                                  f"{r['total_quote']:.2f}", f"{r['avg_quote']:.2f}", f"{r['max_quote']:.2f}", f"{r['min_quote']:.2f}",
                                  f"{r['total_order']:.2f}", f"{r['avg_order']:.2f}", f"{r['total_save']:.2f}", f"{r['avg_save']:.2f}",
                                  r['cost_pct'], f"{r['total_rebate']:.2f}", f"{r['avg_rebate']:.2f}"])
            _write_sheet(ws, '成本分析', headers, data_rows)
            _recolor_header_firstcol(ws, 19)

    # ── 笔记分析 ──
    if note_data:
        from ReportExport.report_generator import (
            build_note_interaction_trend_data, build_note_viral_data,
            build_note_high_value_data, build_note_project_comparison_data,
        )

        analysis_types = note_data.get('analysis_types', [])
        has_content = 'content' in analysis_types
        has_value = 'value' in analysis_types
        has_review = 'review' in analysis_types

        if has_content:
            # Sheet: 内容表现分析
            ws_index += 1
            ws = wb.create_sheet()
            trend_data = build_note_interaction_trend_data(note_data)
            stats = trend_data.get('stats', {})
            if stats:
                headers = ['总互动量', '平均互动量', '中位数互动量', '最大互动量', '平均互动率(%)', '中位数互动率(%)']
                data_rows = [[str(stats.get(k, '-')) for k in
                             ['total_interaction', 'avg_interaction', 'median_interaction', 'max_interaction',
                              'avg_interaction_rate', 'median_interaction_rate']]]
                _write_sheet(ws, '内容表现', headers, data_rows)
                _recolor_header_firstcol_green(ws, 6)

            # Sheet: 爆款笔记
            viral_data = build_note_viral_data(note_data)
            if viral_data:
                ws_index += 1
                ws = wb.create_sheet()
                headers = ['达人昵称', '项目名称', '达人量级', '笔记类型', '互动量', '阅读量', '曝光量', '互动率(%)', 'cpm', 'cpe', '发布时间']
                data_rows = []
                for tier in ['KOC', 'KOL', '十万KOL']:
                    if tier not in viral_data:
                        continue
                    for n in viral_data[tier]:
                        data_rows.append([n['name'], n['project'], n['tier'], n['note_type'], str(n['interaction']),
                                          str(n['read_count']), str(n['exposure']), n['interaction_rate'],
                                          n['cpm'], n['cpe'], n['publish_time'][:10] if n['publish_time'] else ''])
                if data_rows:
                    _write_sheet(ws, '爆款笔记', headers, data_rows)
                    _recolor_header_firstcol_green(ws, 11)

        if has_value:
            ws_index += 1
            ws = wb.create_sheet()
            hv_data = build_note_high_value_data(note_data)
            if hv_data:
                headers = ['达人昵称', '达人量级', '互动量', '成本', 'CPE', '互动率(%)']
                data_rows = []
                for tier in ['KOC', 'KOL', '十万KOL']:
                    if tier not in hv_data:
                        continue
                    for n in hv_data[tier]:
                        data_rows.append([n['name'], n['tier'], str(n['interaction']),
                                          f"{n['cost']:.2f}" if n['cost'] else '0', n['cpe'], n['interaction_rate']])
                if data_rows:
                    _write_sheet(ws, '高价值达人', headers, data_rows)
                    _recolor_header_firstcol_green(ws, 6)

        if has_review:
            ws_index += 1
            ws = wb.create_sheet()
            pc_rows = build_note_project_comparison_data(note_data)
            if pc_rows:
                headers = ['项目名称', '达人量级', '平均互动量', '平均阅读量', '平均成本', '平均CPM', '平均CPE', '平均互动率(%)', '综合得分', '效果评级']
                data_rows = []
                for r in pc_rows:
                    data_rows.append([r['project'], r['tier'], f"{r['avg_interaction']:.0f}", f"{r['avg_read']:.0f}",
                                      f"{r['avg_cost']:.2f}", f"{r['avg_cpm']:.2f}", f"{r['avg_cpe']:.2f}",
                                      f"{r['avg_rate']:.2f}", f"{r['score']:.2f}", r['rating']])
                if data_rows:
                    _write_sheet(ws, '项目效果对比', headers, data_rows)
                    _recolor_header_firstcol_green(ws, 10)

    # 保存到临时文件
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'outputs', 'reports')
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f'审计报告_{week_label}_{ts}.xlsx'
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    logger.info(f"Excel导出成功: {filepath}")
    return filepath, filename
